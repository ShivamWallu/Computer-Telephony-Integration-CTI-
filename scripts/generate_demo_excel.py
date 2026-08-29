import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def create_demo_excel(filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    headers = [
        "Customer ID", "Customer Name", "Mobile Number", "Alternate Number",
        "Email", "Company", "Address", "City", "State", "Pincode",
        "Customer Type", "Status", "Assigned Employee", "Notes"
    ]

    ws.append(headers)

    # Style header row
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    sample_data = [
        ("CUST-1001", "Rajesh Kumar", "+91 98765 43210", "09876543211", "rajesh.kumar@apexlogistics.in", "Apex Global Logistics", "Plot 42, Andheri East", "Mumbai", "Maharashtra", "400001", "VIP", "Active", "Rahul Sharma", "Key logistics client"),
        ("CUST-1002", "Ananya Deshmukh", "9823012345", "+91-98230-12346", "ananya@techvision.io", "TechVision Software Labs", "FC Road, Deccan", "Pune", "Maharashtra", "411004", "Enterprise", "Active", "Amit Verma", "Annual license renewal Q3"),
        ("CUST-1003", "Vikramaditya Rao", "+919811223344", "", "v.rao@zenithenterprises.com", "Zenith Manufacturing Corp", "HITEC City Phase 2", "Hyderabad", "Telangana", "500081", "VIP", "Active", "Priya Patel", "Bulk supply integration"),
        ("CUST-1004", "Sneha Kulkarni", "09845098450", "9845098451", "sneha@bluestarfoods.co.in", "BlueStar Food Processing", "Indiranagar 100ft Rd", "Bengaluru", "Karnataka", "560001", "Standard", "Active", "Rahul Sharma", "Freight forwarding enquiry"),
        ("CUST-1005", "Arjun Malhotra", "+91 98100 55667", "011-26543210", "arjun@malhotragroup.org", "Malhotra Infrastructure Group", "Barakhamba Road, CP", "New Delhi", "Delhi", "110001", "VIP", "Active", "Amit Verma", "Chairman office VIP"),
        ("CUST-1006", "Pooja Hegde", "9830198301", "", "pooja.hegde@orientchemicals.in", "Orient Specialty Chemicals", "Park Street", "Kolkata", "West Bengal", "700016", "Standard", "Lead", "Priya Patel", "Web inbound lead"),
        ("CUST-1007", "Siddharth Jain", "+91-98990-11223", "9899011224", "sjain@silverlinejewels.com", "Silverline Exim Corp", "Johari Bazar", "Jaipur", "Rajasthan", "302001", "VIP", "Active", "Rahul Sharma", "Export compliance"),
        ("CUST-1008", "Meera Nambiar", "9847055443", "+91 98470 55444", "meera@cochinfisheries.com", "Cochin Marine Exports", "Willingdon Island", "Kochi", "Kerala", "682001", "Enterprise", "Active", "Amit Verma", "Cold storage follow-up"),
        ("CUST-1009", "Rohan Gupta", "+91 97112 33445", "", "rohan@dynamicautomotive.in", "Dynamic Auto Components", "Udyog Vihar Phase 4", "Gurugram", "Haryana", "122001", "Standard", "Active", "Priya Patel", "Monthly billing query"),
        ("CUST-1010", "Kavita Reddy", "09885066778", "9885066779", "kavita@apollopharmacies.net", "Sri Krishna Pharma Dist", "MG Road", "Vijayawada", "Andhra Pradesh", "520001", "Enterprise", "Active", "Rahul Sharma", "Pharma distribution"),
        ("CUST-1011", "Manish Tiwari", "+91 94150 99887", "", "m.tiwari@lucknowtextiles.co", "Awadh Handlooms & Textiles", "Hazratganj", "Lucknow", "Uttar Pradesh", "226001", "Standard", "Active", "Amit Verma", "Sample catalog sent"),
        ("CUST-1012", "Divya Sundaram", "9841022334", "044-24567890", "divya@madrasprecision.com", "Madras Precision Tools", "Guindy Industrial Estate", "Chennai", "Tamil Nadu", "600001", "VIP", "Active", "Priya Patel", "Expo contact"),
        ("CUST-1013", "Gaurav Aggarwal", "+91 98140 12345", "", "gaurav@punjabagrohub.in", "Punjab Agro Hub", "GT Road", "Ludhiana", "Punjab", "141001", "Standard", "Lead", "Rahul Sharma", "Automated sorting quote"),
        ("CUST-1014", "Bhavna Patel", "9825066554", "+91 98250 66555", "bhavna@ahmedabadceramics.com", "Morbi Ceramic Exports", "SG Highway", "Ahmedabad", "Gujarat", "380001", "Enterprise", "Active", "Amit Verma", "Export shipment cleared"),
        ("CUST-1015", "Deepak Sen", "+91-94330-77889", "", "deepak@easternpaper.co.in", "Eastern Paper Mills", "GS Road", "Guwahati", "Assam", "781001", "Standard", "Active", "Priya Patel", "Q4 recurring order"),
        ("CUST-1016", "Tarun Chhabra", "09818044332", "9818044333", "tarun@noidaelectronics.in", "NextGen Electronics Ltd", "Sector 62", "Noida", "Uttar Pradesh", "201301", "VIP", "Active", "Rahul Sharma", "Component tracking"),
        ("CUST-1017", "Sunita Menon", "+91 98950 11234", "", "sunita@menonplantations.com", "Malabar Spices & Plantations", "Beach Road", "Kozhikode", "Kerala", "673001", "Standard", "Active", "Amit Verma", "August batch payment"),
        ("CUST-1018", "Harish Bhatt", "9820098200", "+91 98200 98201", "harish@bhattfinance.com", "Bhatt & Associates Advisory", "BKC Complex", "Mumbai", "Maharashtra", "400051", "VIP", "Active", "Priya Patel", "Audit schedule"),
        ("CUST-1019", "Pallavi Shah", "+91 98790 33221", "", "pallavi@surattextilemarket.in", "Surat Weaving Mills", "Ring Road", "Surat", "Gujarat", "395002", "Standard", "Lead", "Rahul Sharma", "Consultation enquiry"),
        ("CUST-1020", "Alok Srivastava", "09450012398", "9450012399", "alok@varanasihospitality.com", "Ganga Heritage Resorts", "Cantt Road", "Varanasi", "Uttar Pradesh", "221001", "Enterprise", "Active", "Amit Verma", "Corporate retreat"),
        ("CUST-1021", "Karthik Raja", "+91-98400-88776", "", "karthik@coimbatorepumps.in", "Kongu Engineering Pumps", "Avinashi Road", "Coimbatore", "Tamil Nadu", "641001", "Standard", "Active", "Priya Patel", "Warranty extension"),
        ("CUST-1022", "Ritu Singhania", "9831099881", "+91 98310 99882", "ritu@singhaniasteel.com", "Singhania Steel & Alloys", "Main Road", "Ranchi", "Jharkhand", "834001", "VIP", "Active", "Rahul Sharma", "Annual contract renewal"),
        ("CUST-1023", "Naveen Choudhary", "+91 94140 55664", "", "naveen@marwarmarbles.com", "Marwar Natural Stones", "Sukher Industrial Area", "Udaipur", "Rajasthan", "313001", "Standard", "Active", "Amit Verma", "Stone supply order"),
        ("CUST-1024", "Shalini Nair", "09847122334", "9847122335", "shalini@keralaspicesonline.com", "Nair Organic Extracts", "Swaraj Round", "Thrissur", "Kerala", "680001", "Enterprise", "Active", "Priya Patel", "EU certification ok"),
        ("CUST-1025", "Amitabh Bose", "+91 98305 44332", "033-22114455", "bose.a@kolkataprint.com", "Bengal Publishing House", "College Street", "Kolkata", "West Bengal", "700001", "VIP", "Active", "Rahul Sharma", "Packaging enquiry")
    ]

    for row_idx, row_data in enumerate(sample_data, start=2):
        ws.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    # Set column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    wb.save(filepath)
    print(f"Generated {filepath} successfully with {len(sample_data)} customer rows.")

if __name__ == "__main__":
    out_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "demo_customers.xlsx")
    create_demo_excel(out_file)
